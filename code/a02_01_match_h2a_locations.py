import marimo

__generated_with = "0.19.11"
app = marimo.App(width="full")


@app.cell
def _():
    import marimo as mo
    from pathlib import Path
    import pyprojroot
    import polars as pl
    from openpyxl import load_workbook
    import numpy as np
    import json
    from itertools import islice
    import re
    import addfips
    import requests
    import urllib
    import time
    DC_STATEHOOD = 1 # Enables DC to be included in the state list
    import us
    import pickle
    import rapidfuzz

    return json, load_workbook, mo, pl, pyprojroot, rapidfuzz, us


@app.cell
def _(pyprojroot):
    root_path = pyprojroot.find_root(criterion='pyproject.toml')
    return (root_path,)


@app.cell
def _(pl):
    # Helper function to read and uppercase all string columns
    def read_census_file(path):
        df = (
            pl.read_csv(path, separator='|', infer_schema=False)
                .with_columns(
                    pl.col(pl.String).str.to_uppercase()
                )
        )
        return df

    return (read_census_file,)


@app.cell
def _(root_path):
    # Load and process Census files
    census_code_path = root_path / 'data' / 'census_geography_codes'
    return (census_code_path,)


@app.cell
def _(census_code_path, pl, read_census_file):
    # Census county names
    census_county = read_census_file(census_code_path / 'national_county2020.txt')
    census_county = (
        census_county
            .with_columns(
                (pl.col("STATEFP") + pl.col("COUNTYFP")).alias('fips')
            )
            .select([
                pl.col("STATE").alias("state"),
                pl.col("COUNTYNAME").alias("county"),
                pl.col("fips")
            ])
            .filter(pl.col("county") != "")
    )
    return (census_county,)


@app.cell
def _(census_code_path, pl, read_census_file):
    # Census counties aggregated to Census places
    census_placebycounty = read_census_file(census_code_path / 'national_place_by_county2020.txt')
    census_place_agg = (
        census_placebycounty
            .with_columns(
                (pl.col("STATEFP") + pl.col("COUNTYFP")).alias('fips')
            )
            .group_by(["STATE", "COUNTYNAME", "PLACENAME"])
            .agg(pl.col("fips").str.join(","))
            .filter(pl.col("PLACENAME") != "")
            .rename({
                "STATE": "state", 
                "COUNTYNAME": "county", 
                "PLACENAME": "place"
            })
    )
    return (census_place_agg,)


@app.cell
def _(census_code_path, pl, read_census_file):
    # Census counties aggregated to ZIP
    census_zip = read_census_file(census_code_path / 'tab20_zcta520_county20_natl.txt')
    census_zip_agg = (
        census_zip
            .group_by("GEOID_ZCTA5_20")
            .agg(pl.col("GEOID_COUNTY_20").str.join(",")
                )
            .filter(pl.col("GEOID_ZCTA5_20") != "")
            .rename({
                "GEOID_ZCTA5_20": "zip",
                "GEOID_COUNTY_20": "fips"
        })
    )
    return (census_zip_agg,)


@app.cell
def _(load_workbook):
    def get_excel_column_names(excel_file_path, sheet_name = None):
        wb = load_workbook(excel_file_path, read_only=True)

        # Select the active sheet if none is specified
        if sheet_name is None:
            sheet = wb.active
        else:
            sheet = wb[sheet_name]

        # Iterate through the first row only to get the headers
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            return list(row)

        return []

    return


@app.cell
def _(root_path):
    h2a_path = root_path / 'data' / 'h2a'
    return (h2a_path,)


@app.cell
def _():
    # Get column names of all the H-2A Excel files
    # Names of H-2A program disclosure files from the DOL-OFLC
    h2a_filenames_dict = {
        2008:'H2A_FY2008.xlsx',
        2009:'H2A_FY2009.xlsx',
        2010:'H-2A_FY2010.xlsx',
        2011:'H-2A_FY2011.xlsx',
        2012:'H-2A_FY2012.xlsx',
        2013:'H2A_FY2013.xls',
        2014:'H-2A_FY14_Q4.xlsx',
        2015:'H-2A_Disclosure_Data_FY15_Q4.xlsx',
        2016:'H-2A_Disclosure_Data_FY16_updated.xlsx',
        2017:'H-2A_Disclosure_Data_FY17.xlsx',
        2018:'H-2A_Disclosure_Data_FY2018_EOY.xlsx',
        2019:'H-2A_Disclosure_Data_FY2019.xlsx',
        2020:'H-2A_Disclosure_Data_FY2020.xlsx',
        2021:'H-2A_Disclosure_Data_FY2021.xlsx',
        2022:'H-2A_Disclosure_Data_FY2022_Q4.xlsx',
        2023:'H-2A_Disclosure_Data_FY2023_Q4.xlsx',
        2024:'H-2A_Disclosure_Data_FY2024_Q4.xlsx'
    }

    # Names of H-2A program Addendum B files from the DOL-OFLC
    add_b_filenames_dict = {
        2020:'H-2A_FY2020_AddendumB_Employment.xlsx',
        2021:'H-2A_Addendum_B_Employment_FY2021.xlsx',
        2022:'H-2A_Addendum_B_Employment_Record_FY2022_Q4.xlsx',
        2023:'H-2A_Addendum_B_Employment_Record_FY2023_Q4.xlsx',
        2024:'H-2A_Addendum_B_Employment_Record_FY2024_Q4.xlsx'
    }
    return add_b_filenames_dict, h2a_filenames_dict


@app.cell
def _():
    # h2a_dtype_dict = {}

    # h2a_dtype_dict['2008'] = {
    #     'CASE_NO':'string',
    #     'CASE_STATUS':'string',
    #     'EMPLOYER_NAME':'string',
    #     'EMPLOYER_CITY':'string',
    #     'EMPLOYER_STATE':'string',
    #     'EMPLOYER_POSTAL_CODE':'string',
    #     'NBR_WORKERS_CERTIFIED':'string',
    #     'CERTIFICATION_BEGIN_DATE':'string',
    #     'CERTIFICATION_END_DATE':'string',
    #     'BASIC_RATE_OF_PAY':'string',
    #     'BASIC_UNIT_OF_PAY':'string',
    #     'ALIEN_WORK_CITY':'string',
    #     'ALIEN_WORK_STATE':'string',
    #     'ORGANIZATION_FLAG':'string'
    # }

    # h2a_dtype_dict['2009'] = {
    #     'CASE_NO':'string',
    #     'CASE_STATUS':'string',
    #     'EMPLOYER_NAME':'string',
    #     'EMPLOYER_CITY':'string',
    #     'EMPLOYER_STATE':'string',
    #     'EMPLOYER_POSTAL_CODE':'string',
    #     'NBR_WORKERS_CERTIFIED':'string',
    #     'CERTIFICATION_BEGIN_DATE':'string',
    #     'CERTIFICATION_END_DATE':'string',
    #     'BASIC_RATE_OF_PAY':'string',
    #     'BASIC_UNIT_OF_PAY':'string',
    #     'ALIEN_WORK_CITY':'string',
    #     'ALIEN_WORK_STATE':'string',
    #     'ORGANIZATION_FLAG':'string'
    # }

    # h2a_dtype_dict['2010'] = h2a_dtype_dict['2009']

    # h2a_dtype_dict['2011'] = {
    #     'CASE_NO':'string',
    #     'CASE_STATUS':'string',
    #     'EMPLOYER_NAME':'string',
    #     'EMPLOYER_CITY':'string',
    #     'EMPLOYER_STATE':'string',
    #     'EMPLOYER_POSTAL_CODE':'string',
    #     'NBR_WORKERS_REQUESTED':'string',
    #     'NBR_WORKERS_CERTIFIED':'string',
    #     'REQUESTED_START_DATE_OF_NEED':'object',
    #     'REQUESTED_END_DATE_OF_NEED':'object',
    #     'CERTIFICATION_BEGIN_DATE':'string',
    #     'CERTIFICATION_END_DATE':'string',
    #     'BASIC_NUMBER_OF_HOURS':'string',
    #     'BASIC_RATE_OF_PAY':'string',
    #     'BASIC_UNIT_OF_PAY':'string',
    #     'ALIEN_WORK_CITY':'string',
    #     'ALIEN_WORK_STATE':'string',
    #     'ORGANIZATION_FLAG':'string'
    # }

    # h2a_dtype_dict['2012'] = h2a_dtype_dict['2011']

    # h2a_dtype_dict['2013'] = {
    #     'CASE_NO':'string',
    #     'CASE_STATUS':'string',
    #     'EMPLOYER_NAME':'string',
    #     'EMPLOYER_CITY':'string',
    #     'EMPLOYER_STATE':'string',
    #     'EMPLOYER_POSTAL_CODE':'string',
    #     'NBR_WORKERS_CERTIFIED':'string',
    #     'REQUESTED_START_DATE_OF_NEED':'object',
    #     'REQUESTED_END_DATE_OF_NEED':'object',
    #     'CERTIFICATION_BEGIN_DATE':'string',
    #     'CERTIFICATION_END_DATE':'string',
    #     'BASIC_NUMBER_OF_HOURS':'string',
    #     'BASIC_RATE_OF_PAY':'string',
    #     'BASIC_UNIT_OF_PAY':'string',
    #     'ALIEN_WORK_CITY':'string',
    #     'ALIEN_WORK_STATE':'string',
    #     'ORGANIZATION_FLAG':'string'
    # }

    # h2a_dtype_dict['2014'] = {
    #     'CASE_NO':'string',
    #     'CASE_STATUS':'string',
    #     'EMPLOYER_NAME':'string',
    #     'EMPLOYER_CITY':'string',
    #     'EMPLOYER_STATE':'string',
    #     'EMPLOYER_POSTAL_CODE':'string',
    #     'NBR_WORKERS_CERTIFIED':'string',
    #     'REQUESTED_START_DATE_OF_NEED':'object',
    #     'REQUESTED_END_DATE_OF_NEED':'object',
    #     'CERTIFICATION_BEGIN_DATE':'string',
    #     'CERTIFICATION_END_DATE':'string',
    #     'BASIC_NUMBER_OF_HOURS':'string',
    #     'BASIC_RATE_OF_PAY':'string',
    #     'BASIC_UNIT_OF_PAY':'string',
    #     'WORKSITE_LOCATION_CITY':'string',
    #     'WORKSITE_LOCATION_STATE':'string',
    #     'ORGANIZATION_FLAG':'string'
    # }

    # h2a_dtype_dict['2015'] = {
    #     'CASE_NUMBER':'string',
    #     'CASE_STATUS':'string',
    #     'EMPLOYER_NAME':'string',
    #     'EMPLOYER_CITY':'string',
    #     'EMPLOYER_STATE':'string',
    #     'EMPLOYER_POSTAL_CODE':'string',
    #     'NBR_WORKERS_REQUESTED':'string',
    #     'NBR_WORKERS_CERTIFIED':'string',
    #     'CERTIFICATION_BEGIN_DATE':'string',
    #     'CERTIFICATION_END_DATE':'string',
    #     'BASIC_NUMBER_OF_HOURS':'string',
    #     'BASIC_RATE_OF_PAY':'string',
    #     'BASIC_UNIT_OF_PAY':'string',
    #     'WORKSITE_CITY':'string',
    #     'WORKSITE_STATE':'string',
    #     'WORKSITE_POSTAL_CODE':'string',
    #     'ORGANIZATION_FLAG':'string'
    # }

    # h2a_dtype_dict['2016'] = {
    #     'CASE_NUMBER':'string',
    #     'CASE_STATUS':'string',
    #     'EMPLOYER_NAME':'string',
    #     'EMPLOYER_CITY':'string',
    #     'EMPLOYER_STATE':'string',
    #     'EMPLOYER_POSTAL_CODE':'string',
    #     'NBR_WORKERS_REQUESTED':'string',
    #     'NBR_WORKERS_CERTIFIED':'string',
    #     'REQUESTED_START_DATE_OF_NEED':'object',
    #     'REQUESTED_END_DATE_OF_NEED':'object',
    #     'JOB_START_DATE':'string',
    #     'JOB_END_DATE':'string',
    #     'BASIC_NUMBER_OF_HOURS':'string',
    #     'BASIC_RATE_OF_PAY':'string',
    #     'BASIC_UNIT_OF_PAY':'string',
    #     'WORKSITE_CITY':'string',
    #     'WORKSITE_STATE':'string',
    #     'WORKSITE_POSTAL_CODE':'string',
    #     'ORGANIZATION_FLAG':'string',
    #     'PRIMARY/SUB':'string',
    # }

    # h2a_dtype_dict['2017'] = {
    #     'CASE_NUMBER':'string',
    #     'CASE_STATUS':'string',
    #     'EMPLOYER_NAME':'string',
    #     'EMPLOYER_CITY':'string',
    #     'EMPLOYER_STATE':'string',
    #     'EMPLOYER_POSTAL_CODE':'string',
    #     'NBR_WORKERS_REQUESTED':'string',
    #     'NBR_WORKERS_CERTIFIED':'string',
    #     'REQUESTED_START_DATE_OF_NEED':'object',
    #     'REQUESTED_END_DATE_OF_NEED':'object',
    #     'JOB_START_DATE':'string', 
    #     'JOB_END_DATE':'string',
    #     'BASIC_NUMBER_OF_HOURS':'string',
    #     'BASIC_RATE_OF_PAY':'string',
    #     'BASIC_UNIT_OF_PAY':'string',
    #     'WORKSITE_CITY':'string',
    #     'WORKSITE_COUNTY':'string',
    #     'WORKSITE_STATE':'string',
    #     'WORKSITE_POSTAL_CODE':'string',
    #     'ORGANIZATION_FLAG':'string',
    #     'PRIMARY/SUB':'string'
    # }

    # h2a_dtype_dict['2018'] = {
    #     'CASE_NO':'string',
    #     'CASE_STATUS':'string',
    #     'EMPLOYER_NAME':'string',
    #     'EMPLOYER_CITY':'string',
    #     'EMPLOYER_STATE':'string',
    #     'EMPLOYER_POSTAL_CODE':'string',
    #     'NBR_WORKERS_REQUESTED':'string',
    #     'NBR_WORKERS_CERTIFIED':'string',
    #     'REQUESTED_START_DATE_OF_NEED':'object',
    #     'REQUESTED_END_DATE_OF_NEED':'object',
    #     'JOB_START_DATE':'string', 
    #     'JOB_END_DATE':'string',
    #     'BASIC_NUMBER_OF_HOURS':'string',
    #     'BASIC_RATE_OF_PAY':'string',
    #     'BASIC_UNIT_OF_PAY':'string',
    #     'WORKSITE_CITY':'string',
    #     'WORKSITE_COUNTY':'string',
    #     'WORKSITE_STATE':'string',
    #     'WORKSITE_POSTAL_CODE':'string',
    #     'ORGANIZATION_FLAG':'string',
    #     'PRIMARY_SUB':'string'
    # }

    # h2a_dtype_dict['2019'] = {
    #     'CASE_NUMBER':'string',
    #     'CASE_STATUS':'string',
    #     'EMPLOYER_NAME':'string',
    #     'EMPLOYER_CITY':'string',
    #     'EMPLOYER_STATE':'string',
    #     'EMPLOYER_POSTAL_CODE':'string',
    #     'NBR_WORKERS_REQUESTED':'string',
    #     'NBR_WORKERS_CERTIFIED':'string',
    #     'REQUESTED_START_DATE_OF_NEED':'object',
    #     'REQUESTED_END_DATE_OF_NEED':'object',
    #     'JOB_START_DATE':'string', 
    #     'JOB_END_DATE':'string',
    #     'BASIC_NUMBER_OF_HOURS':'string',
    #     'BASIC_RATE_OF_PAY':'string',
    #     'BASIC_UNIT_OF_PAY':'string',
    #     'WORKSITE_CITY':'string',
    #     'WORKSITE_COUNTY':'string',
    #     'WORKSITE_STATE':'string',
    #     'WORKSITE_POSTAL_CODE':'string',
    #     'ORGANIZATION_FLAG':'string',
    #     'PRMARY/SUB':'string'
    # }

    # h2a_dtype_dict['2020'] = {
    #     'CASE_NUMBER':'string',
    #     'CASE_STATUS':'string',
    #     'EMPLOYER_NAME':'string',
    #     'EMPLOYER_CITY':'string',
    #     'EMPLOYER_STATE':'string',
    #     'EMPLOYER_POSTAL_CODE':'string',
    #     'TOTAL_WORKERS_NEEDED':'string',
    #     'TOTAL_WORKERS_H2A_REQUESTED':'string',
    #     'TOTAL_WORKERS_H2A_CERTIFIED':'string',
    #     'REQUESTED_BEGIN_DATE':'string',
    #     'REQUESTED_END_DATE':'string',
    #     'EMPLOYMENT_BEGIN_DATE':'string',
    #     'EMPLOYMENT_END_DATE':'string',
    #     'ANTICIPATED_NUMBER_OF_HOURS':'string',
    #     'WAGE_OFFER':'string',
    #     'PER':'string',
    #     'WORKSITE_CITY':'string',
    #     'WORKSITE_COUNTY':'string',
    #     'WORKSITE_STATE':'string',
    #     'WORKSITE_POSTAL_CODE':'string',
    #     'TYPE_OF_EMPLOYER_APPLICATION':'string',
    #     'H2A_LABOR_CONTRACTOR':'string'
    # }

    # h2a_dtype_dict['2021'] = h2a_dtype_dict['2020']
    # h2a_dtype_dict['2022'] = h2a_dtype_dict['2020']
    # h2a_dtype_dict['2023'] = h2a_dtype_dict['2020']

    # h2a_dtype_dict['2024'] = {
    #     'CASE_NUMBER':'string',
    #     'CASE_STATUS':'string',
    #     'EMPLOYER_NAME':'string',
    #     'EMPLOYER_CITY':'string',
    #     'EMPLOYER_STATE':'string',
    #     'EMPLOYER_POSTAL_CODE':'string',
    #     'TOTAL_WORKERS_NEEDED':'string',
    #     'TOTAL_WORKERS_H2A_REQUESTED':'string',
    #     'TOTAL_WORKERS_H2A_CERTIFIED':'string',
    #     'REQUESTED_BEGIN_DATE':'string',
    #     'REQUESTED_END_DATE':'string',
    #     'EMPLOYMENT_BEGIN_DATE':'string',
    #     'EMPLOYMENT_END_DATE':'string',
    #     'ANTICIPATED_NUMBER_OF_HOURS':'string',
    #     'WAGE_OFFER':'string',
    #     'PER':'string',
    #     'WORKSITE_CITY':'string',
    #     'WORKSITE_COUNTY':'string',
    #     'WORKSITE_STATE':'string',
    #     'WORKSITE_POSTAL_CODE':'string',
    #     'TYPE_OF_EMPLOYER_APPLICATION':'string',
    #     'AG_ASSN_OR_AGENCY_STATUS':'string',
    #     'H2A_LABOR_CONTRACTOR':'string'
    # }

    # def analyze_schema_diffs(dtype_dict):
    #     years = sorted(dtype_dict.keys())

    #     for i in range(len(years)):
    #         curr_year = years[i]
    #         curr_schema = dtype_dict[curr_year]

    #         print(f"\n--- {curr_year} ---")

    #         if i == 0:
    #             print(f"Initial State: {len(curr_schema)} columns")
    #             continue

    #         prev_year = years[i-1]
    #         prev_schema = dtype_dict[prev_year]

    #         # Check if they are identical
    #         if curr_schema == prev_schema:
    #             print(f"Identical to {prev_year}")
    #             continue

    #         # Find differences
    #         curr_keys = set(curr_schema.keys())
    #         prev_keys = set(prev_schema.keys())

    #         added = curr_keys - prev_keys
    #         removed = prev_keys - curr_keys
    #         common = curr_keys & prev_keys
    #         changed_type = {k for k in common if curr_schema[k] != prev_schema[k]}

    #         if added: print(f"  Added:   {sorted(list(added))}")
    #         if removed: print(f"  Removed: {sorted(list(removed))}")
    #         if changed_type:
    #             for k in changed_type:
    #                 print(f"Type Change: {k} ({prev_schema[k]} -> {curr_schema[k]})")

    # analyze_schema_diffs(h2a_dtype_dict)
    return


@app.cell
def _(pl):
    # Define the set of columns we want to add and remove over the years
    def mutate_list(base, add, remove):
        final = base + add
        if remove:
            final = [s for s in final if s not in remove]
        return final

    def s_map(cols):
        s_dict = {}
        for c in cols:
            if 'DATE' in c:
                s_dict[c] = pl.Date
            else:
                s_dict[c] = pl.String
        return s_dict

    add_cols_dict = {}
    remove_cols_dict = {}

    add_cols_dict[2011] = [
        'BASIC_NUMBER_OF_HOURS',
        'NBR_WORKERS_REQUESTED',
        'REQUESTED_END_DATE_OF_NEED',
        'REQUESTED_START_DATE_OF_NEED'
    ]
    remove_cols_dict[2013] = [
        'NBR_WORKERS_REQUESTED'
    ]
    add_cols_dict[2014] = [
        'WORKSITE_LOCATION_CITY',
        'WORKSITE_LOCATION_STATE'
    ]
    remove_cols_dict[2014] =[
        'ALIEN_WORK_CITY',
        'ALIEN_WORK_STATE'
    ]
    add_cols_dict[2015] = [
        'CASE_NUMBER',
        'NBR_WORKERS_REQUESTED',
        'WORKSITE_CITY',
        'WORKSITE_POSTAL_CODE',
        'WORKSITE_STATE'
    ]
    remove_cols_dict[2015] = [
        'CASE_NO',
        'REQUESTED_END_DATE_OF_NEED',
        'REQUESTED_START_DATE_OF_NEED',
        'WORKSITE_LOCATION_CITY',
        'WORKSITE_LOCATION_STATE'
    ]
    add_cols_dict[2016] = [
        'JOB_END_DATE',
        'JOB_START_DATE',
        'PRIMARY/SUB',
        'REQUESTED_END_DATE_OF_NEED',
        'REQUESTED_START_DATE_OF_NEED'
    ]
    remove_cols_dict[2016] = [
        'CERTIFICATION_BEGIN_DATE',
        'CERTIFICATION_END_DATE'
    ]
    add_cols_dict[2017] = [
        'WORKSITE_COUNTY'
    ]
    add_cols_dict[2018] = [
        'CASE_NO',
        'PRIMARY_SUB'
    ]
    remove_cols_dict[2018] = [
        'CASE_NUMBER',
        'PRIMARY/SUB'
    ]
    add_cols_dict[2019] = [
        'CASE_NUMBER',
        'PRMARY/SUB'
    ]
    remove_cols_dict[2019] = [
        'CASE_NO',
        'PRIMARY_SUB'
    ]
    add_cols_dict[2020] = [
        'ANTICIPATED_NUMBER_OF_HOURS',
        'EMPLOYMENT_BEGIN_DATE',
        'EMPLOYMENT_END_DATE',
        'H2A_LABOR_CONTRACTOR',
        'PER',
        'REQUESTED_BEGIN_DATE',
        'REQUESTED_END_DATE',
        'TOTAL_WORKERS_H2A_CERTIFIED',
        'TOTAL_WORKERS_H2A_REQUESTED',
        'TOTAL_WORKERS_NEEDED',
        'TYPE_OF_EMPLOYER_APPLICATION',
        'WAGE_OFFER'
    ]
    remove_cols_dict[2020] = [
        'BASIC_NUMBER_OF_HOURS',
        'BASIC_RATE_OF_PAY',
        'BASIC_UNIT_OF_PAY',
        'JOB_END_DATE',
        'JOB_START_DATE',
        'NBR_WORKERS_CERTIFIED',
        'NBR_WORKERS_REQUESTED',
        'ORGANIZATION_FLAG',
        'PRMARY/SUB',
        'REQUESTED_END_DATE_OF_NEED',
        'REQUESTED_START_DATE_OF_NEED'
    ]
    add_cols_dict[2024] = [
        'AG_ASSN_OR_AGENCY_STATUS'
    ]
    return add_cols_dict, mutate_list, remove_cols_dict, s_map


@app.cell
def _(add_cols_dict, mutate_list, remove_cols_dict, s_map):
    # Generate set of schemas
    cols_dict = {}
    schema_dict = {}

    cols_dict[2008] = [
        'CASE_NO',
        'CASE_STATUS',
        'EMPLOYER_NAME',
        'EMPLOYER_CITY',
        'EMPLOYER_STATE',
        'EMPLOYER_POSTAL_CODE',
        'NBR_WORKERS_CERTIFIED',
        'CERTIFICATION_BEGIN_DATE',
        'CERTIFICATION_END_DATE',
        'BASIC_RATE_OF_PAY',
        'BASIC_UNIT_OF_PAY',
        'ALIEN_WORK_CITY',
        'ALIEN_WORK_STATE',
        'ORGANIZATION_FLAG'
    ]
    schema_dict[2008] = s_map(cols_dict[2008])

    for _y in range(2009, 2025):
        add_list = add_cols_dict.get(_y, [])
        remove_list = remove_cols_dict.get(_y, [])

        cols_dict[_y] = mutate_list(cols_dict[_y-1], add_list, remove_list)  
        schema_dict[_y] = s_map(cols_dict[_y])
    return cols_dict, schema_dict


@app.cell
def _(s_map):
    # Columns and schema from the Addendum B files
    add_b_cols_dict = {}
    add_b_schema_dict = {}

    add_b_cols_dict[2020] = [
        'CASE_NUMBER',
        'NAME_OF_AGRICULTURAL_BUSINESS',
        'PLACE_OF_EMPLOYMENT_CITY',
        'PLACE_OF_EMPLOYMENT_STATE',
        'PLACE_OF_EMPLOYMENT_POSTAL_CODE',
        'TOTAL_WORKERS'
    ]
    add_b_cols_dict[2021] = add_b_cols_dict[2020]
    add_b_cols_dict[2022] = add_b_cols_dict[2020]
    add_b_cols_dict[2023] = add_b_cols_dict[2020]
    add_b_cols_dict[2024] = add_b_cols_dict[2020]

    for _y, _cols in add_b_cols_dict.items():
        add_b_schema_dict[_y] = s_map(_cols)
    return add_b_cols_dict, add_b_schema_dict


@app.cell
def _():
    # Define common column names
    h2a_rename_dict = {
        'CASE_NO':'case_number',
        'CASE_NUMBER':'case_number',
        'CASE_STATUS':'case_status',
        'EMPLOYER_NAME':'employer_name',
        'EMPLOYER_CITY':'employer_city',
        'EMPLOYER_STATE':'employer_state',
        'EMPLOYER_POSTAL_CODE':'employer_postal_code',
        'NBR_WORKERS_REQUESTED':'nbr_workers_requested',
        'NBR_WORKERS_CERTIFIED':'nbr_workers_certified',
        'TOTAL_WORKERS_NEEDED':'nbr_workers_needed',
        'TOTAL_WORKERS_H2A_REQUESTED':'nbr_workers_requested',
        'TOTAL_WORKERS_H2A_CERTIFIED':'nbr_workers_certified',
        'REQUESTED_START_DATE_OF_NEED':'requested_begin_date',
        'REQUESTED_END_DATE_OF_NEED':'requested_end_date',
        'CERTIFICATION_BEGIN_DATE':'certification_begin_date',
        'CERTIFICATION_END_DATE':'certification_end_date',
        'REQUESTED_BEGIN_DATE':'requested_begin_date',
        'REQUESTED_END_DATE':'requested_end_date',
        'EMPLOYMENT_BEGIN_DATE':'job_begin_date',
        'EMPLOYMENT_END_DATE':'job_end_date',
        'JOB_START_DATE':'job_begin_date',
        'JOB_END_DATE':'job_end_date',
        'BASIC_NUMBER_OF_HOURS':'number_of_hours',
        'ANTICIPATED_NUMBER_OF_HOURS':'number_of_hours',
        'BASIC_RATE_OF_PAY':'wage_rate',
        'WAGE_OFFER':'wage_rate',
        'BASIC_UNIT_OF_PAY':'wage_unit',
        'PER':'wage_unit',
        'ALIEN_WORK_CITY':'worksite_city',
        'ALIEN_WORK_STATE':'worksite_state',
        'WORKSITE_LOCATION_CITY':'worksite_city',
        'WORKSITE_LOCATION_STATE':'worksite_state',
        'WORKSITE_CITY':'worksite_city',
        'WORKSITE_COUNTY':'worksite_county',
        'WORKSITE_STATE':'worksite_state',
        'WORKSITE_POSTAL_CODE':'worksite_zip',
        'PRIMARY/SUB':'primary_sub',
        'PRIMARY_SUB':'primary_sub',
        'PRMARY/SUB':'primary_sub',
        'ORGANIZATION_FLAG':'organization_flag',
        'TYPE_OF_EMPLOYER_APPLICATION':'type_of_employer_application',
        'H2A_LABOR_CONTRACTOR':'h2a_labor_contractor',
        'AG_ASSN_OR_AGENCY_STATUS':'ag_association_or_agency',
    }

    # Common names in Addendum B files
    add_b_rename_dict = {
        'CASE_NUMBER':'case_number',
        'NAME_OF_AGRICULTURAL_BUSINESS':'business_name',
        'PLACE_OF_EMPLOYMENT_CITY':'worksite_city',
        'PLACE_OF_EMPLOYMENT_STATE':'worksite_state',
        'PLACE_OF_EMPLOYMENT_POSTAL_CODE':'worksite_zip',
        'TOTAL_WORKERS':'total_h2a_workers_requested'
    }
    return add_b_rename_dict, h2a_rename_dict


@app.cell
def _(
    add_b_cols_dict,
    add_b_filenames_dict,
    add_b_rename_dict,
    add_b_schema_dict,
    cols_dict,
    h2a_filenames_dict,
    h2a_path,
    h2a_rename_dict,
    pl,
    schema_dict,
):
    # Read the H-2A performance data files
    h2a_df = pl.DataFrame()
    for _y in range(2008, 2025):
        _path = h2a_path / h2a_filenames_dict[_y]
        print(f'Reading H-2A for year {_y}')
        _df = pl.read_excel(
            source=_path,
            columns=cols_dict[_y],
            schema_overrides=schema_dict[_y],
            engine='calamine'
        )

        _df = _df.with_columns(
            pl.col(pl.Date).dt.to_string(),
            pl.col(pl.String).fill_null(""),
            fiscal_year=_y
        ).rename(
            mapping=h2a_rename_dict, strict=False
        )

        h2a_df = pl.concat(items=[h2a_df, _df], how='diagonal')

    # Same with Addendum B files
    # Read the H-2A performance data files
    add_b_df = pl.DataFrame()
    for _y in range(2020, 2025):
        _path = h2a_path / add_b_filenames_dict[_y]
        print(f'Reading Addendum B for year {_y}')
        _df = pl.read_excel(
            source=_path,
            columns=add_b_cols_dict[_y],
            schema_overrides=add_b_schema_dict[_y],
            engine='calamine'
        )

        _df = _df.with_columns(
            pl.col(pl.Date).dt.to_string(),
            pl.col(pl.String).fill_null(""),
            fiscal_year=_y
        ).rename(
            mapping=add_b_rename_dict, strict=False
        )

        add_b_df = pl.concat(items=[add_b_df, _df], how='diagonal')
    return add_b_df, h2a_df


@app.cell
def _():
    # # Function that reads lists of H-2A disclosure files and returns a dict of all dataframes
    # def read_h2a_excel_files(year_filename_dict, year_list, year_dtype_dict):
    #     h2a_df_dict = {}
    #     for y in year_list:
    #         year = str(y)
    #         filename = year_filename_dict[year]
    #         print(filename)
    #         dtype_dict = year_dtype_dict[year]
    #         col_list = list(dtype_dict.keys())
    #         h2a_df_dict[year] = pd.read_excel('../Data/h2a/' + filename, dtype=dtype_dict, usecols=col_list, parse_dates=False, na_filter=False)

    #     return h2a_df_dict

    # # Process H-2A files
    # years_to_load = list(range(2008, 2025))
    # _h2a_df_dict_read = read_h2a_excel_files(h2a_filenames_dict, years_to_load, h2a_dtype_dict)

    # # Pickle
    # with open("json/h2a_pickle", "wb") as _fp:
    #     pickle.dump(_h2a_df_dict_read, _fp)

    # # Process Addendum B files
    # add_b_years_to_load = list(range(2020, 2025))
    # _add_b_df_dict_read = read_h2a_excel_files(add_b_filenames_dict, add_b_years_to_load, add_b_dtype_dict)

    # # Pickle
    # with open("json/h2a_add_b_pickle", "wb") as _fp:
    #     pickle.dump(_add_b_df_dict_read, _fp)

    # with open("json/h2a_pickle", "rb") as _fp:
    #     h2a_df_dict = pickle.load(_fp)

    # with open("json/h2a_add_b_pickle", "rb") as _fp:
    #     add_b_df_dict = pickle.load(_fp)

    # # Define function that concatenates all years together
    # def concatenate_h2a_years(h2a_df_dict, h2a_rename_dict):
    #     h2a_all_years_df = pd.DataFrame()
    #     for year, df in h2a_df_dict.items():
    #         df = df.rename(columns=h2a_rename_dict)
    #         df['fiscal_year'] = int(year)

    #         # Concatenate
    #         h2a_all_years_df = pd.concat([h2a_all_years_df, df], ignore_index=True, sort=False)

    #     # Standardize NAs, convert all to string, remove time from dates
    #     h2a_df = h2a_all_years_df.fillna(value='')
    #     h2a_df = h2a_df.astype("string")

    #     for columns in h2a_df.columns:
    #         h2a_df[columns] = h2a_df[columns].str.replace(' 00:00:00', '', regex=False)
    #         h2a_df[columns] = h2a_df[columns].str.upper()

    #     return h2a_df

    # h2a_df = concatenate_h2a_years(h2a_df_dict, h2a_rename_dict)
    # add_b_df = concatenate_h2a_years(add_b_df_dict, add_b_rename_dict)
    return


@app.cell
def _(mo):
    mo.md(r"""
    Add FIPS codes using worksite location information
    """)
    return


@app.cell
def _(add_b_df, h2a_df):
    # Set of unique worksite locations we need to find counties for
    h2a_worksite_locations = h2a_df[['worksite_city', 'worksite_county', 'worksite_state', 'worksite_zip']]
    add_b_worksite_locations = add_b_df[['worksite_city', 'worksite_state', 'worksite_zip']]
    h2a_worksite_locations = h2a_worksite_locations.drop_duplicates()
    add_b_worksite_locations = add_b_worksite_locations.drop_duplicates()
    return add_b_worksite_locations, h2a_worksite_locations


@app.cell
def _(mo):
    mo.md(r"""
    Given input location information (city, county, state, ZIP), we want to write a function that finds FIPS codes
    """)
    return


@app.cell
def _():
    # Write a function that fixes typos in county name and explodes county names
    def clean_county_explode(df, county_col, state_col):

        # Replace separators with commas
        for sep in [' AND ', ' & ', '/']:
            df[county_col] = df[county_col].str.replace(sep, ',')

        # Explode
        df['county_list'] = df[county_col].str.split(',')
        df_exploded = df.explode('county_list').reset_index(drop=True)

        # Remove whitespacess
        df_exploded['county_list'] = df_exploded['county_list'].str.strip()

        # Strip suffixes from names
        for suffix in [' COUNTY', ' COUNTIES', ' PARISH', ' PARRISH']:
            df_exploded['county_list'] = df_exploded['county_list'].str.replace(suffix, '')

        # Fix common typos for Lousiana counties
        la_typo_dict = {
            'ST\\.\\w':'ST. ',
            'NORTH\\. ':'NORTH ',
            'SOUTH\\. ':'SOUTH ',
            'EAST\\. ':'EAST ',
            'WEST\\. ':'WEST ',
            'BATON ROGUE':'BATON ROUGE',
            'JEFF DAVIS':'JEFFERSON DAVIS',
            'IBERIAL':'IBERIA'
        }

        for typo, fix in la_typo_dict.items():
            df_exploded.loc[df_exploded[state_col] == 'LA', 'county_list'] = df_exploded[df_exploded[state_col] == 'LA']['county_list'].str.replace(typo, fix, regex=True)

        return df_exploded

    # Cleans ZIP codes, mostly by removing trailing 4 digits for ZIP+4 codes
    def clean_zip(df, zip_col):

        # Remove period from ZIP codes
        df[zip_col] = df[zip_col].str.replace('.', '', regex=False)

        # Remove the 4 extra trailing digits after hyphen from ZIP codes
        df[zip_col] = df[zip_col].str.split('-').str[0]

        # Pad with 0s from the left
        # df['zip'] = df['zip'].str.pad(width=5, side='left', fillchar='0')

        return df

    return clean_county_explode, clean_zip


@app.function
# Function that add FIPS using addfips
def add_fips_using_addfips(df, county_col, state_col, name_of_new_fips_col):
    import addfips
    af = addfips.AddFIPS()
    df[name_of_new_fips_col] = df.apply(lambda x: af.get_county_fips(x[county_col], state=x[state_col]), axis=1)
    df = df.fillna(value='')

    return df


@app.function
# A function that add FIPS using ZIP and census ZIP to county crosswalk; also does a sanity check by matching states
def add_fips_using_census_zip(df, zip_col, state_col, name_of_new_fips_col, zip_fips_df, check_df):
    zip_to_fips_dict = dict(zip(zip_fips_df['zip'], zip_fips_df['fips']))

    # Match ZIP codes to FIPS
    df[name_of_new_fips_col] = df[zip_col].map(zip_to_fips_dict)
    df = df.fillna(value='')

    # Sanity check by matching states
    # fips_to_state_dict = dict(zip(check_df['fips'], check_df['state']))
    # df['state_check'] = df[name_of_new_fips_col].map(fips_to_state_dict)

    # Null out FIPS where state does not match
    # df.loc[df['state_check'] != df[state_col], name_of_new_fips_col] = ''
    # Note that this also nulls out zipcodes that match to multiple county FIPS

    # df = df.drop(columns=['state_check'])
    return df


@app.cell
def _(rapidfuzz):
    # Function for fuzzy string matching
    def fuzz_search(df, df_match_col, df_state_col, df_fips_col, state, name_to_match):

        # Search only within the same state
        state_df = df[df[df_state_col] == state]

        # Strip place suffixes from names
        place_suffixes = [' CITY', ' TOWN', ' VILLAGE', ' CDP', ' MUNICIPALITY', ' BOROUGH', ' TOWNSHIP', ' CENSUS AREA', ' CENSUS DESIGNATED PLACE', ' COUNTY', ' PARISH']
        for suffix in place_suffixes:
            state_df[df_match_col] = state_df[df_match_col].str.replace(suffix, '', regex=False)

        # Get match score
        state_df['score'] = state_df[df_match_col].apply(lambda _x: rapidfuzz.fuzz.partial_ratio_alignment(_x, name_to_match, processor = rapidfuzz.utils.default_process).score)
        state_df = state_df.sort_values('score')

        # Get highest scoring match
        state_df = state_df.sort_values('score')
        max_score_row = state_df[state_df['score'] == state_df['score'].max()].reset_index()

        # Check if there is a match
        if len(max_score_row) >= 1:
            fips = str(max_score_row[df_fips_col][0])
            score = str(max_score_row['score'][0])
            census_name = max_score_row[df_match_col][0]
            return (fips, score, census_name)  # Best match
        else:
            return ('', '', '')

    return (fuzz_search,)


@app.cell
def _(us):
    # Function cleans state names using the US package and returns 2 character abbreviations; returns empty string if no match; DC is included
    def clean_state_to_abbr(input):
        if input == '':
            return ''

        # Include DC
        if str(input).upper() == 'DISTRICT OF COLUMBIA':
            return 'DC'
        if str(input).upper() == 'DC':
            return 'DC'

        # US package lookup can handle names, abbreviations, FIPs
        state_obj = us.states.lookup(str(input).strip())

        return state_obj.abbr if state_obj else ''

    return (clean_state_to_abbr,)


@app.cell
def _(
    census_county,
    census_place_agg,
    clean_county_explode,
    clean_state_to_abbr,
    clean_zip,
    fuzz_search,
    pd,
):
    # Define wrapper function that takes input location dataframe and returns dataframe with possible matched FIPS codes
    def assign_best_fips(df, county_col, city_col, state_col, zip_col, zip_fips_df, state_fips_df, place_fips_df, fuzzy_score):

        # Locations are defined by worksite_city, worksite_county, worksite_state, worksite_zip
        # We will be operating on a new set of columns instead to preserve the original location identifiers
        df['xcity'] = df[city_col]
        df['xcounty'] = df[county_col]
        df['xstate'] = df[state_col].apply(lambda _x: clean_state_to_abbr(_x))
        df['xzip'] = df[zip_col]

        # Explode by counties, clean ZIP codes
        df_exploded = clean_county_explode(df, 'xcounty', 'xstate')
        df_exploded = clean_zip(df_exploded, 'xzip')

        # Add FIPS with addfips
        df_exploded_with_addfips = add_fips_using_addfips(df_exploded, 'county_list', 'xstate', 'fips_from_addfips')

        # Add FIPS with census ZIP to county crosswalk
        df_exploded_with_addfips_zip = add_fips_using_census_zip(df_exploded_with_addfips, 'xzip', 'xstate', 'fips_from_census_zip', zip_fips_df, state_fips_df)

        # Use fuzzy string matching for remaining unmatched rows
        # Grab the columns we want to do fuzzy matching on
        df_exploded_addfips_zip_unmatched = df_exploded_with_addfips_zip[(df_exploded_with_addfips_zip['fips_from_addfips'] == '') & (df_exploded_with_addfips_zip['fips_from_census_zip'] == '')]
        df_exploded_addfips_zip_unmatched = df_exploded_addfips_zip_unmatched[['xcity', 'county_list', 'xstate', 'xzip']]

        # Perform fuzzy string matching for county names
        county_name_fuzzy_match_df = df_exploded_addfips_zip_unmatched.apply(
        lambda _x: fuzz_search(
            census_county,
            'county',
            'state',
            'fips',
            _x['xstate'],
            _x['county_list']
        ),
        axis=1,
        result_type='expand'
        )

        county_name_fuzzy_match_df = county_name_fuzzy_match_df[[0, 1, 2]]
        county_name_fuzzy_match_df = county_name_fuzzy_match_df.rename(columns={0:'fips_from_fuzzy_county', 1:'fuzzy_county_score', 2:'fuzzy_county_name'})

        # Merge fuzzy matched county names back to unmatched dataframe
        df_exploded_addfips_zip_unmatched = df_exploded_addfips_zip_unmatched.merge(county_name_fuzzy_match_df, left_index=True, right_index=True, how='left')

        # Perform fuzzy string matching for city names
        city_name_fuzzy_match_df = df_exploded_addfips_zip_unmatched.apply(
        lambda _x: fuzz_search(
            census_place_agg,
            'place',
            'state',
            'fips',
            _x['xstate'],
            _x['xcity']
        ),
        axis=1,
        result_type='expand'
        )

        city_name_fuzzy_match_df = city_name_fuzzy_match_df[[0, 1, 2]]
        city_name_fuzzy_match_df = city_name_fuzzy_match_df.rename(columns={0:'fips_from_fuzzy_city', 1:'fuzzy_city_score', 2:'fuzzy_city_name'})

        # Merge fuzzy matched city names back to unmatched dataframe
        df_exploded_addfips_zip_unmatched = df_exploded_addfips_zip_unmatched.merge(city_name_fuzzy_match_df, left_index=True, right_index=True, how='left')

        # New England has to be matched again as they put place names in their county field
        ne_states = ['CT', 'ME', 'MA', 'NH', 'RI', 'VT']
        ne_fuzzy_match_df = df_exploded_addfips_zip_unmatched.apply(
            lambda _x: fuzz_search(
                census_place_agg,
                'place',
                'state',
                'fips',
                _x['xstate'],
                _x['county_list']
            ),
            axis=1,
            result_type='expand'
            )

        ne_fuzzy_match_df = ne_fuzzy_match_df[[0, 1, 2]]
        ne_fuzzy_match_df = ne_fuzzy_match_df.rename(columns={0:'fips_from_fuzzy_ne', 1:'fuzzy_ne_score', 2:'fuzzy_ne_name'})

        # Add NE fuzzy matches to unmatched dataframe
        df_exploded_addfips_zip_unmatched = df_exploded_addfips_zip_unmatched.merge(ne_fuzzy_match_df, left_index=True, right_index=True, how='left')

        # Add all 3 fuzzy matches back to original matched dataframe
        df_exploded_with_addfips_zip_fuzzy = df_exploded_with_addfips_zip.merge(df_exploded_addfips_zip_unmatched, on = ['xcity', 'county_list', 'xstate', 'xzip'], how='left')

        # Convert score columns to numeric
        score_columns = ['fuzzy_county_score', 'fuzzy_city_score', 'fuzzy_ne_score']
        for col in score_columns:
            df_exploded_with_addfips_zip_fuzzy[col] = pd.to_numeric(df_exploded_with_addfips_zip_fuzzy[col], errors='coerce')

        # Apply FIPS selection and validation function to rows of matched FIPS
        df_exploded_with_addfips_zip_fuzzy['final_fips'] = df_exploded_with_addfips_zip_fuzzy.apply(
            lambda _x: pick_best_fips(
                _x,
                'fips_from_addfips',
                'fips_from_census_zip',
                'fips_from_fuzzy_county',
                'fuzzy_county_score',
                'fips_from_fuzzy_city',
                'fuzzy_city_score',
                'fips_from_fuzzy_ne',
                'fuzzy_ne_score',
                'xstate',
                80
            ),
            axis=1
        )

        return df_exploded_with_addfips_zip_fuzzy

    return (assign_best_fips,)


@app.function
# Function that validates and picks the best FIPS code for each row
def pick_best_fips(row, county_name_fips, zip_fips, fuzzy_county_name_fips, fuzzy_county_name_score, fuzzy_city_name_fips, fuzzy_city_name_score, fuzzy_ne_name_fips, fuzzy_ne_name_score, state_col, fuzzy_score):
    # Priority order:
    #1. Census ZIP to county crosswalk
    if row[zip_fips] != '':
        return row[zip_fips]
    #2. addfips county name match
    elif row[county_name_fips] != '':
        return row[county_name_fips]
    # Special case: New England states - if fuzzy NE score is high, use it
    if row[state_col] in ['CT', 'ME', 'MA', 'NH', 'RI', 'VT'] and row[fuzzy_ne_name_fips] != '' and (row[fuzzy_ne_name_score] > fuzzy_score + 10):
        return row[fuzzy_ne_name_fips]
    #3. Fuzzy county name and fuzzy city name both have high scores and agree
    elif (row[fuzzy_county_name_fips] != '') and (row[fuzzy_city_name_fips] != '') and (row[fuzzy_county_name_fips] == row[fuzzy_city_name_fips]) and (row[fuzzy_county_name_score] > fuzzy_score) and (row[fuzzy_city_name_score] > fuzzy_score):
        return row[fuzzy_county_name_fips]
    #4. Fuzzy county name and fuzzy city name both have high scores but disagree
    elif (row[fuzzy_county_name_fips] != '') and (row[fuzzy_city_name_fips] != '') and (row[fuzzy_county_name_fips] != row[fuzzy_city_name_fips]) and (row[fuzzy_county_name_score] > fuzzy_score) and (row[fuzzy_city_name_score] > fuzzy_score):
        # Pick the one with higher score with higher threshold
        if row[fuzzy_county_name_score] >= row[fuzzy_city_name_score] and (row[fuzzy_county_name_score] > (fuzzy_score + 5)):
            return row[fuzzy_county_name_fips]
        elif row[fuzzy_county_name_score] < row[fuzzy_city_name_score] and (row[fuzzy_city_name_score] > (fuzzy_score + 5)):
            return row[fuzzy_city_name_fips]
    #5. Fuzzy county name match only
    elif row[fuzzy_county_name_fips] != '' and (row[fuzzy_county_name_score] > fuzzy_score + 10):
        return row[fuzzy_county_name_fips]
    #6. Fuzzy city name match only
    elif row[fuzzy_city_name_fips] != '' and (row[fuzzy_city_name_score] > fuzzy_score + 10):
        return row[fuzzy_city_name_fips]
    else:
        return ''


@app.cell
def _(
    assign_best_fips,
    census_county,
    census_place_agg,
    census_zip_agg,
    h2a_worksite_locations,
):
    # Match FIPS for H-2A worksites
    h2a_worksite_locations_added_fips = assign_best_fips(h2a_worksite_locations, 'worksite_county', 'worksite_city', 'worksite_state', 'worksite_zip', census_zip_agg, census_county, census_place_agg, 80)
    return (h2a_worksite_locations_added_fips,)


@app.cell
def _(
    add_b_worksite_locations,
    assign_best_fips,
    census_county,
    census_place_agg,
    census_zip_agg,
):
    # For Addendum B, we need to add the non-existent county column
    add_b_worksite_locations['worksite_county'] = ''
    add_b_worksite_locations_added_fips = assign_best_fips(add_b_worksite_locations, 'worksite_county', 'worksite_city', 'worksite_state', 'worksite_zip', census_zip_agg, census_county, census_place_agg, 80)
    return (add_b_worksite_locations_added_fips,)


@app.cell
def _(h2a_worksite_locations_added_fips):
    h2a_worksite_locations_added_fips[h2a_worksite_locations_added_fips['final_fips'] == '']
    return


@app.cell
def _(add_b_worksite_locations_added_fips):
    add_b_worksite_locations_added_fips[add_b_worksite_locations_added_fips['final_fips'] == '']
    return


@app.cell
def _(add_b_worksite_locations_added_fips, h2a_worksite_locations_added_fips):
    # Export both sets of unmatched locations
    def export_unmatched_locations(df):
        unmatched_df = df[df['final_fips'] == '']
        unmatched_df = unmatched_df[['xcity', 'xstate', 'xzip', 'county_list']]
        unmatched_df = unmatched_df.drop_duplicates()
        unmatched_df = unmatched_df.rename(columns = {'xcity':'city', 'xstate':'state', 'xzip':'zip', 'county_list':'county'}).reset_index(drop=True)
        return unmatched_df

    export_df = export_unmatched_locations(h2a_worksite_locations_added_fips)
    export_df.to_csv("json/unmatched_h2a_locations.csv", index=False)

    export_df = export_unmatched_locations(add_b_worksite_locations_added_fips)
    export_df.to_csv("json/unmatched_add_b_locations.csv", index=False)
    return


@app.cell
def _(mo):
    mo.md(r"""
    # Use Gemini + Places API to get matches for the unmatched locations, then come back here and use them to fill in the missing FIPS
    """)
    return


@app.cell
def _(json, pd):
    # Load mappings we obtained
    with open('json/placeid_address_components_mapping_json.json') as _fp:
        placeid_address_components_map = json.load(_fp)

    h2a_placeid = pd.read_parquet("../binaries/h2a_location_placeids.parquet")
    add_b_placeid = pd.read_parquet("../binaries/add_b_location_placeids.parquet")
    return add_b_placeid, h2a_placeid, placeid_address_components_map


@app.cell
def _(json):
    # Function that takes a list of Place IDs and returns a list of counties
    def placeid_to_county(placeids, placeid_address_components_map):
        county_list = []
        for placeid in placeids:
            if placeid not in placeid_address_components_map:
                print(f'Place ID {placeid} not in mapping')
                continue

            responses = json.loads(placeid_address_components_map[placeid])['addressComponents']
            for response in responses:
                if not 'types' in response: # We are not interested in street address components
                    continue
                response_type = response['types']

                if 'locality' in response_type and 'political' in response_type:
                    locality = response['longText']
                elif 'administrative_area_level_2' in response_type and 'political' in response_type:
                    county = response['longText']
                    county_list.append(county)
                elif 'administrative_area_level_1' in response_type and 'political' in response_type:
                    state = response['longText']
                else:
                    continue

        return county_list

    return (placeid_to_county,)


@app.cell
def _(
    add_b_placeid,
    h2a_placeid,
    placeid_address_components_map,
    placeid_to_county,
):
    h2a_placeid['original_county'] = h2a_placeid.apply(lambda _x: placeid_to_county(_x['original_location_placeid'], placeid_address_components_map), axis=1)
    h2a_placeid['suggested_county'] = h2a_placeid.apply(lambda _x: placeid_to_county(_x['suggested_location_placeid'], placeid_address_components_map), axis=1)

    add_b_placeid['original_county'] = add_b_placeid.apply(lambda _x: placeid_to_county(_x['original_location_placeid'], placeid_address_components_map), axis=1)
    add_b_placeid['suggested_county'] = add_b_placeid.apply(lambda _x: placeid_to_county(_x['suggested_location_placeid'], placeid_address_components_map), axis=1)

    # Function produces common counties returned via Places API calls on both original and Gemini-suggested location names
    def common_counties(list1, list2):
        return list(set(list1).intersection(list2))

    def common_state(state1, state2):
        if state1 == state2:
            return state1
        else:
            return ''

    h2a_placeid['common_county'] = h2a_placeid.apply(lambda _x: common_counties(_x['original_county'], _x['suggested_county']), axis = 1)
    # h2a_placeid['common_state'] = h2a_placeid.apply(lambda _x: common_state(_x['state_name'], _x['state_name_suggested']), axis = 1)
    add_b_placeid['common_county'] = add_b_placeid.apply(lambda _x: common_counties(_x['original_county'], _x['suggested_county']), axis = 1)
    # add_b_placeid['common_state'] = add_b_placeid.apply(lambda _x: common_state(_x['state_name'], _x['state_name_suggested']), axis = 1)
    return


@app.cell
def _(add_b_placeid, h2a_placeid):
    # Function that add FIPS using addfips
    def add_fips_to_places_api_county(county_list, state):
        import addfips
        af = addfips.AddFIPS()
        string_of_new_fips = ''
        for county_name in county_list:
            county_fips = af.get_county_fips(county_name, state)
            if not county_fips:
                county_fips = ''
            string_of_new_fips = string_of_new_fips + county_fips + ','

        string_of_new_fips = string_of_new_fips.rstrip(',')
        return string_of_new_fips

    h2a_placeid['common_county_fips'] = h2a_placeid.apply(lambda _x: add_fips_to_places_api_county(_x['common_county'], _x['state_name']), axis=1)
    add_b_placeid['common_county_fips'] = add_b_placeid.apply(lambda _x: add_fips_to_places_api_county(_x['common_county'], _x['state_name']), axis=1)
    return


@app.cell
def _(add_b_placeid, h2a_placeid):
    # Combine FIPs from Places API back to original dataset
    # Change name back to column names from original dataset
    h2a_places_api_fips = h2a_placeid.rename(columns = {
        'city': 'xcity',
        'county': 'county_list',
        'state': 'xstate',
        'zip': 'xzip',
        'common_county_fips': 'places_api_fips'
    })
    h2a_places_api_fips = h2a_places_api_fips[['xcity', 'xstate', 'xzip', 'county_list', 'places_api_fips']]

    add_b_places_api_fips = add_b_placeid.rename(columns = {
        'city': 'xcity',
        'county': 'county_list',
        'state': 'xstate',
        'zip': 'xzip',
        'common_county_fips': 'places_api_fips'
    })
    add_b_places_api_fips = add_b_places_api_fips[['xcity', 'xstate', 'xzip', 'county_list', 'places_api_fips']]
    return add_b_places_api_fips, h2a_places_api_fips


@app.cell
def _(
    add_b_places_api_fips,
    add_b_worksite_locations_added_fips,
    h2a_places_api_fips,
    h2a_worksite_locations_added_fips,
    pd,
):
    # Define new finalized FIPs, keep only columns we need, then collapse back into form before I exploded xcounty into county_list
    # Location are defined by the worksite columns, the ones prefixed with x are just copies of worksite columns that I modify for explosion
    h2a_worksite_locations_finalized = pd.merge(h2a_worksite_locations_added_fips, h2a_places_api_fips, how='left' ,on = ['xcity', 'xstate', 'county_list', 'xzip']).fillna('')
    h2a_worksite_locations_finalized = h2a_worksite_locations_finalized[['worksite_city', 'worksite_county', 'worksite_state', 'worksite_zip', 'final_fips', 'places_api_fips']]

    add_b_worksite_locations_finalized = pd.merge(add_b_worksite_locations_added_fips, add_b_places_api_fips, how='left' ,on = ['xcity', 'xstate', 'county_list', 'xzip']).fillna('')
    add_b_worksite_locations_finalized = add_b_worksite_locations_finalized[['worksite_city', 'worksite_county', 'worksite_state', 'worksite_zip', 'final_fips', 'places_api_fips']]

    # Finalized FIPS
    def fips_selector(final_fips, places_api_fips):
        if final_fips != '':
            return final_fips
        elif places_api_fips:
            return places_api_fips
        else:
            return ''

    h2a_worksite_locations_finalized['fips'] = h2a_worksite_locations_finalized.apply(lambda _x: fips_selector(_x['final_fips'], _x['places_api_fips']), axis=1)
    add_b_worksite_locations_finalized['fips'] = add_b_worksite_locations_finalized.apply(lambda _x: fips_selector(_x['final_fips'], _x['places_api_fips']), axis=1)

    #h2a_worksite_locations_finalized = h2a_worksite_locations_finalized.drop(columns = ['final_fips', 'places_api_fips'])
    #add_b_worksite_locations_finalized = add_b_worksite_locations_finalized.drop(columns = ['final_fips', 'places_api_fips'])
    return add_b_worksite_locations_finalized, h2a_worksite_locations_finalized


@app.cell
def _(add_b_worksite_locations_finalized, h2a_worksite_locations_finalized):
    h2a_worksite_locations_collapsed = h2a_worksite_locations_finalized.groupby(['worksite_city', 'worksite_county', 'worksite_state', 'worksite_zip']).agg({'fips': ','.join})
    add_b_worksite_locations_collapsed = add_b_worksite_locations_finalized.groupby(['worksite_city', 'worksite_state', 'worksite_zip']).agg({'fips': ','.join})
    return add_b_worksite_locations_collapsed, h2a_worksite_locations_collapsed


@app.cell
def _(
    add_b_df,
    add_b_worksite_locations_collapsed,
    h2a_df,
    h2a_worksite_locations_collapsed,
    pd,
):
    h2a_with_fips_df = pd.merge(h2a_df, h2a_worksite_locations_collapsed, on=['worksite_city', 'worksite_county', 'worksite_state', 'worksite_zip'], how='left')
    add_b_with_fips_df = pd.merge(add_b_df, add_b_worksite_locations_collapsed, on=['worksite_city', 'worksite_state', 'worksite_zip'], how='left')
    return add_b_with_fips_df, h2a_with_fips_df


@app.cell
def _(add_b_with_fips_df, h2a_with_fips_df):
    # Export binary
    h2a_with_fips_df.to_parquet("../binaries/h2a_with_fips.parquet", index=False)
    add_b_with_fips_df.to_parquet("../binaries/h2a_addendum_b_with_fips.parquet", index=False)
    return


@app.cell
def _(h2a_with_fips_df):
    h2a_with_fips_df[h2a_with_fips_df['nbr_workers_needed'] != '']
    return


@app.cell
def _():
    return


if __name__ == "__main__":
    app.run()
